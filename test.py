import time
from concurrent.futures import ThreadPoolExecutor
import tools
from params import Scenario_name
from llm_agent import LLM_Agent
from memory import DrivingMemory
from matplotlib.animation import FuncAnimation
import matplotlib.animation as animation
import os
from time import gmtime, strftime
from idm_controller import IDM
import matplotlib.pyplot as plt
from bayesian_game_agent import Bayesian_Agent
import random
import numpy as np
import openpyxl
import re
from vehicle import Pedestrian

# Dynamically load the correct environment to access pedestrian splines
if Scenario_name == 'intersection':
    from scenario_environment import intersection_environment as environment
elif Scenario_name == 'merge':
    from scenario_environment import merge_environment as environment
elif Scenario_name == 'roundabout':
    from scenario_environment import roundabout_environment as environment
else:
    raise ValueError('no such environment, check Scenario_name in params')

Sim_times = 200
suffix = '(with-instruction&HMI)'

def open_excel(i):
    file_dir = './test/' + Scenario_name + '/excel/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
    file_name = file_dir + str(i) + '.xlsx'

    if not os.path.exists(file_dir):
        os.makedirs(file_dir)
    workbook = openpyxl.Workbook()
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    return file_name, workbook


def clean_val(val):
    """Remove invalid control characters that cause openpyxl IllegalCharacterError."""
    if isinstance(val, str):
        return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F]', '', val)
    return val


def write_data(workbook, vehicles, llm_output, if_passed, t):
    column_names = ['t', 'x', 'y', 'v', 'acc', 'theta', 'dis2des', 'type', 'action', 'HDVintent', 'HDVstyle', 'HMI', 'if_passed']
    
    clean_llm = [clean_val(out) for out in llm_output]

    for vehicle in vehicles:
        sheet_name = str(vehicle.id)
        if sheet_name not in workbook.sheetnames:
            worksheet = workbook.create_sheet(sheet_name)
            worksheet.append(column_names)
        else:
            worksheet = workbook[sheet_name]
            
        # Handle the difference between Vehicles and Pedestrians for logging
        behavior_or_agg = getattr(vehicle, 'behavior', getattr(vehicle, 'aggressiveness', 'N/A'))
        v_type = getattr(vehicle, 'type', 'vehicle')

        state = [
            round(vehicle.x, 2), 
            round(vehicle.y, 2), 
            round(vehicle.speed, 2), 
            round(vehicle.acc, 2), 
            round(vehicle.heading, 2), 
            round(vehicle.dis2des, 2), 
            behavior_or_agg, 
            v_type,
            clean_llm[0], 
            clean_llm[2], 
            clean_llm[3], 
            clean_llm[1], 
            if_passed
        ]
        
        row_data = [t] + state

        row_data = [clean_val(item) for item in row_data]
        state = [clean_val(item) for item in state]

        worksheet.append(row_data)
        worksheet.cell(row=t + 2, column=1, value=t)
        for i, item in enumerate(state):
            worksheet.cell(row=t + 2, column=i + 2, value=item)
            
    return workbook


class Simulator:
    def __init__(self, case_id, seed):
        self.seed = seed
        random.seed(self.seed)
        np.random.seed(self.seed)
        
        # Spawn the Ego car and Surrounding HDVs
        self.cav_info, self.hdv_infos = tools.initialize_vehicles(num=3) 
        
        # --- NEW: Spawn a Pedestrian ---
        # Choose a crosswalk based on the environment
        if Scenario_name == 'merge':
            loc = random.choice(['s', 'm'])
        else:
            loc = random.choice(['n', 's', 'e', 'w'])
            
        direction = random.choice([1, -1])
        ped = Pedestrian(loc, direction, 'normal', ped_id=99)
        ped.x, ped.y, ped.speed, ped.heading, ped.dis2des, ped.max_speed = environment.default_pedestrian_state(loc, direction)
        
        # Inject the pedestrian into the tracking radar
        self.hdv_infos.append(ped)
        
        self.case_id = case_id
        self.agent = LLM_Agent()
        self.memory = DrivingMemory()
        self.fig, self.ax = plt.subplots(figsize=(8, 8))
        self.instruction_info = None
        self.retrieved_instruction_info = None
        self.llm_output = [None, None, None, None]  
        self.stop_threads = False  
        self.st = time.time()
        self.executor = ThreadPoolExecutor(max_workers=2)
        self.file_name, self.workbook = open_excel(case_id)

    def run(self):
        " ---- option 1: show animation ---- "
        # ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        # plt.show()

        " ---- option 2: save as gif ---- "
        ani = FuncAnimation(self.fig, self.update, interval=10, frames=Sim_times, blit=False, repeat=False, save_count=Sim_times)
        video_dir = './test/' + Scenario_name + '/video/' + strftime("%Y-%m-%d", gmtime()) + suffix + '/'
        if not os.path.exists(video_dir):
            os.makedirs(video_dir)
        ani.save(video_dir + str(self.case_id) + '.gif', dpi=50)
        print('saved')
        plt.close()

    def update(self, frame):
        print('This is frame:', frame, round(time.time() - self.st, 2))
        time_now = time.time()
        if frame >= Sim_times - 1:
            self.stop_threads = True
            print('shutting down')

        primary_hdv = tools.find_opponent(self.cav_info, self.hdv_infos)

        # Skip conflict checks if the primary target is a pedestrian to avoid TTC math crashes
        if getattr(primary_hdv, 'type', 'vehicle') != 'pedestrian' and tools.if_passed_conflict_point(self.cav_info, primary_hdv):
            self.llm_output[0] = 'FASTER'
            print('no conflict anymore')
            self.stop_threads = True
        else:
            self.actor()
            self.executor.submit(self.reasoner)

        llm_action, hmi_advice = self.llm_output[0], self.llm_output[1]
        
        # Determine acceleration based on the primary threat
        if getattr(primary_hdv, 'type', 'vehicle') == 'pedestrian':
            # Hardcoded yielding IDM for pedestrians
            if llm_action in ['SLOWER', 'IDLE']:
                ego_acc = -2.0 if self.cav_info.speed > 0 else 0.0
            else:
                ego_acc = 1.0 # Fallback if LLM hallucinations
        else:
            controller = IDM(self.cav_info, primary_hdv, llm_action)
            ego_acc = controller.cal_acceleration()
            
        temp_cav_info = tools.kinematic_model(self.cav_info, ego_acc)
        
        temp_hdv_infos = []
        for entity in self.hdv_infos:
            if getattr(entity, 'type', 'vehicle') == 'pedestrian':
                # Pedestrians walk at a constant pace
                temp_hdv_infos.append(tools.kinematic_model(entity, 0.0))
            else:
                bayesian_agent = Bayesian_Agent(entity, self.cav_info, action_type='discrete')
                temp_hdv_infos.append(bayesian_agent.update_state())
            
        self.hdv_infos = temp_hdv_infos
        self.cav_info = temp_cav_info
        
        tools.plot_figs(self.cav_info, self.hdv_infos, self.ax, self.llm_output, self.instruction_info, self.retrieved_instruction_info)
        
        if getattr(primary_hdv, 'type', 'vehicle') == 'pedestrian':
            if_passed = False
        else:
            if_passed = tools.if_passed_conflict_point(self.cav_info, primary_hdv)
            
        self.instruction_info = tools.generate_simulation_hdv_instruction(self.cav_info, primary_hdv)
        
        workbook = write_data(self.workbook, self.hdv_infos + [self.cav_info], self.llm_output, if_passed, frame)
        workbook.save(self.file_name)
        
        time_end = time.time()
        if time_end - time_now < 0.1:
            time.sleep(0.1 - time_end + time_now)

    def actor(self):
        sce_descrip = tools.scenario_experience_generator(self.cav_info, self.hdv_infos, self.llm_output, self.instruction_info)
        retrieved_memory = self.memory.retrieveMemory(query_scenario=sce_descrip, top_k=1)
        time_start = time.time()
        retrieve_time = round(time.time() - time_start, 5)
        print('Retrieve memory time/s', retrieve_time)
        self.llm_output[0] = retrieved_memory[0][0]['final_action']
        self.retrieved_instruction_info = retrieved_memory[1][0]
        self.ax.text(20, -90, f'Fast retrieve time: {round(retrieve_time, 2)}')

    def reasoner(self):
        time_start = time.time()
        if self.stop_threads:
            return
        output = self.agent.llm_run(self.llm_output, self.instruction_info, self.cav_info, self.hdv_infos, self.memory, if_train_mode=False)
        self.llm_output[1:] = output[1:]
        interference_time = round(time.time() - time_start, 2)
        print('Interference time/s', interference_time)
        self.ax.text(20, -80, f'Slow interference time: {round(interference_time, 2)}')


case_num = 50
SEED_TABLE = [_ for _ in range(case_num)]
for case in range(case_num):
    seed = SEED_TABLE[case]
    sim = Simulator(case, seed)
    sim.run()